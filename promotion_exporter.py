import re
"""
Módulo genérico para exportação de promoções por marketplace
Suporta múltiplos canais (Shopee, Mercado Livre, etc) com templates específicos
Garante fidelidade absoluta ao template de saída do marketplace
"""

import pandas as pd
import numpy as np
import unicodedata
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class PromotionExporter:
    """Exporta promoções no formato compatível com diferentes marketplaces."""
    
    # Mapeamento de sinônimos de colunas de ENTRADA (para identificar dados)
    COLUMN_SYNONYMS = {
        "id": ["sku", "mlb", "id do produto", "id_produto", "product_id", "item_id", "codigo", "product id", "item id", "sku ou mlb"],
        "descricao": ["descricao", "titulo", "nome do produto", "nome_produto", "product_name", "name", "product name"],
        "preco": ["preco", "price", "valor", "valor_venda", "preco_venda", "valor venda", "preco venda", "preco sugerido", "preço sugerido", "preco limite", "preço limite", "preco promo limite", "preço promo limite", "preco atual", "preço atual", "preco atual r", "preço atual r"]
    }
    
    # Templates de SAÍDA por marketplace (estrutura exata que será exportada)
    MARKETPLACE_TEMPLATES = {
        "Shopee": {
            "colunas_saida": [
                "ID do produto",
                "Nome do Produto. (Opcional)",
                "Nº de Ref. Parent SKU. (Opcional)",
                "ID de variação",
                "Variação de nome. (Opcional)",
                "Nº de Ref. SKU. (Opcional)",
                "Preço original (opcional)",
                "Preço de desconto",
                "Limite de compra (Opcional)",
            ]
        },
        "Mercado Livre": {
            "colunas_saida": [
                "TITLE",
                "ITEM_ID",
                "SKU",
                "ORIGINAL_PRICE",
                "DISCOUNT_PERCENTAGE",
                "FINAL_PRICE",
                "SUGGESTION",
                "RECEIVES",
                "LOYALTY_DISCOUNT_PERCENTAGE",
                "LOYALTY_PRICE",
                "LOYALTY_RECEIVES",
                "STATUS",
                "ACTION",
                "ERRORS"
            ],
            "tem_cabecalhos_customizados": True,
            "cabecalho_linha_1": [
                "Título do anúncio ",
                "Número do anúncio ",
                "SKU ",
                "Preço original ",
                "Desconto ",
                "Desconto ",
                "Desconto ",
                "Desconto ",
                "Desconto exclusivo Meli+ \n Ofereça um desconto para assinantes do Meli+, eles compram 78% a mais do que outras pessoas. É opcional.",
                "Desconto exclusivo Meli+ ",
                "Desconto exclusivo Meli+ ",
                "Status da promoção ",
                "O que você quer fazer com este anúncio? ",
                "Observações e erros \nNão altere este campo"
            ],
            "cabecalho_linha_2": [
                None,
                None,
                None,
                None,
                "Porcentagem",
                "Preço final",
                "Avaliação do desconto",
                "Você recebe",
                "Porcentagem",
                "Preço final",
                "Você recebe",
                None,
                None,
                None
            ]
        }
    }
    
    def __init__(self, marketplace="Shopee"):
        """
        Inicializa o exportador
        
        Args:
            marketplace: Nome do marketplace (ex: "Shopee", "Mercado Livre")
        """
        if marketplace not in self.MARKETPLACE_TEMPLATES:
            raise ValueError(f"Marketplace '{marketplace}' não suportado. Suportados: {list(self.MARKETPLACE_TEMPLATES.keys())}")
        
        self.marketplace = marketplace
        self.template = self.MARKETPLACE_TEMPLATES[marketplace]
    
    def _normalizar_nome_coluna(self, nome):
        """
        Normaliza o nome de uma coluna para comparação
        Remove espaços, acentos, caracteres especiais e converte para minúsculas
        """
        # Remover acentos
        nome = ''.join(
            c for c in unicodedata.normalize('NFD', str(nome))
            if unicodedata.category(c) != 'Mn'
        )
        # Converter para minúsculas
        nome = nome.lower()
        # Remover tudo que não for letra ou número
        nome = re.sub(r'[^a-z0-9]', '', nome)
        return nome.strip()
    
    def _encontrar_coluna(self, df, tipo_coluna):
        """
        Encontra uma coluna no DataFrame usando sinônimos
        
        Args:
            df: DataFrame
            tipo_coluna: "id", "descricao" ou "preco"
            
        Returns:
            Nome da coluna encontrada ou None
        """
        sinonimos = self.COLUMN_SYNONYMS.get(tipo_coluna, [])
        
        # Normalizar sinônimos
        sinonimos_normalizados = [self._normalizar_nome_coluna(s) for s in sinonimos]
        
        # Procurar nas colunas do DataFrame
        for col in df.columns:
            col_normalizado = self._normalizar_nome_coluna(col)
            if col_normalizado in sinonimos_normalizados:
                return col
        
        return None
    
    def _normalizar_dataframe(self, df):
        """
        Normaliza o DataFrame identificando e renomeando colunas automaticamente
        
        Args:
            df: DataFrame com colunas variadas
            
        Returns:
            DataFrame com colunas padronizadas (id_original, descricao_original, preco_original)
        """
        df_norm = df.copy()
        
        # Encontrar e renomear colunas
        col_id = self._encontrar_coluna(df_norm, "id")
        col_descricao = self._encontrar_coluna(df_norm, "descricao")
        col_preco = self._encontrar_coluna(df_norm, "preco")
        
        # Validar se encontrou as colunas essenciais
        if not col_id or not col_descricao or not col_preco:
            colunas_faltantes = []
            if not col_id:
                colunas_faltantes.append("ID (SKU/MLB/ID do Produto)")
            if not col_descricao:
                colunas_faltantes.append("Descrição (Título/Nome)")
            if not col_preco:
                colunas_faltantes.append("Preço")
            
            raise ValueError(f"Não foi possível identificar as colunas: {', '.join(colunas_faltantes)}. Colunas disponíveis: {list(df_norm.columns)}")
        
        # Armazenar os nomes originais das colunas encontradas
        df_norm["_id_original"] = df_norm[col_id].astype(str)
        df_norm["_descricao_original"] = df_norm[col_descricao].astype(str)
        df_norm["_preco_original"] = pd.to_numeric(df_norm[col_preco], errors='coerce')
        
        return df_norm
    
    def filtrar_por_categoria(self, df, categoria="oportunidade", margem_minima=15.0, margem_alvo=30.0):
        """
        Filtra produtos por categoria (Curva ABC ou Oportunidades)
        Sincronizado com a lógica do Dashboard principal
        
        Args:
            df: DataFrame com dados de produtos
            categoria: "oportunidade", "curva_a", "curva_b", "curva_c", "saudavel", "alerta", "prejuizo"
            margem_minima: Margem minima (DEPRECADO - usar margem_alvo + 5%)
            margem_alvo: Margem alvo configurada no dashboard (padrao 30%)
            
        Returns:
            DataFrame filtrado
        """
        df_filtrado = df.copy()
        
        # Debug: Verificar colunas disponíveis
        colunas_disponiveis = df_filtrado.columns.tolist()
        
        if categoria.lower() == "oportunidade":
            # Produtos de oportunidade (Curva B/C com margem 5% acima da alvo e saudaveis)
            # EXATAMENTE como no Dashboard (linhas 1601-1603 do app.py)
            
            # Verificar se as colunas necessarias existem
            if "Curva ABC" not in colunas_disponiveis or "Status" not in colunas_disponiveis:
                return pd.DataFrame(columns=df_filtrado.columns)
            
            # Aplicar EXATAMENTE a mesma logica do Dashboard
            df_filtrado = df_filtrado[
                (df_filtrado["Curva ABC"].astype(str).str.contains("B", na=False) | 
                 df_filtrado["Curva ABC"].astype(str).str.contains("C", na=False)) &
                (df_filtrado["Status"] == "🟢 Saudavel")
            ]
        
        elif categoria.lower() == "curva_a":
            # Apenas Curva A
            if "Curva ABC" in colunas_disponiveis:
                df_filtrado = df_filtrado[df_filtrado["Curva ABC"].astype(str).str.contains("A", na=False, case=False)]
            else:
                return pd.DataFrame(columns=df_filtrado.columns)
        
        elif categoria.lower() == "curva_b":
            # Apenas Curva B
            if "Curva ABC" in colunas_disponiveis:
                df_filtrado = df_filtrado[df_filtrado["Curva ABC"].astype(str).str.contains("B", na=False, case=False)]
            else:
                return pd.DataFrame(columns=df_filtrado.columns)
        
        elif categoria.lower() == "curva_c":
            # Apenas Curva C
            if "Curva ABC" in colunas_disponiveis:
                df_filtrado = df_filtrado[df_filtrado["Curva ABC"].astype(str).str.contains("C", na=False, case=False)]
            else:
                return pd.DataFrame(columns=df_filtrado.columns)
        
        elif categoria.lower() == "saudavel":
            # Produtos saudáveis
            if "Status" in colunas_disponiveis:
                status_col = df_filtrado["Status"].astype(str)
                df_filtrado = df_filtrado[
                    status_col.str.contains("Saudável", na=False, case=False) | \
                    status_col.str.contains("Saudavel", na=False, case=False)
                ]
            else:
                return pd.DataFrame(columns=df_filtrado.columns)
        
        elif categoria.lower() == "alerta":
            # Produtos em alerta
            if "Status" in colunas_disponiveis:
                df_filtrado = df_filtrado[df_filtrado["Status"].astype(str).str.contains("Alerta", na=False, case=False)]
            else:
                return pd.DataFrame(columns=df_filtrado.columns)
        
        elif categoria.lower() == "prejuizo":
            # Produtos em prejuízo
            if "Status" in colunas_disponiveis:
                df_filtrado = df_filtrado[df_filtrado["Status"].astype(str).str.contains("Prejuízo|Prejuizo", na=False, case=False)]
            else:
                return pd.DataFrame(columns=df_filtrado.columns)
        
        return df_filtrado.reset_index(drop=True)
    
    def mapear_dados_para_marketplace(self, df, desconto_percent=0.0):
        """
        Mapeia dados do dashboard para o formato EXATO do marketplace
        
        Args:
            df: DataFrame com dados de produtos
            desconto_percent: Percentual de desconto a aplicar (ex: 0.05 para 5%)
            
        Returns:
            DataFrame formatado exatamente como o template do marketplace
        """
        # Normalizar DataFrame para identificar colunas automaticamente
        df_norm = self._normalizar_dataframe(df)
        
        # Calcular preço com desconto
        df_norm["_preco_desconto"] = (df_norm["_preco_original"] * (1 - desconto_percent)).round(2)
        
        # Mapear para Mercado Livre
        if self.marketplace == "Mercado Livre":
            df_ml = pd.DataFrame()
            
            for col in self.template["colunas_saida"]:
                if col == "TITLE":
                    df_ml[col] = df_norm["_descricao_original"]
                elif col == "ITEM_ID":
                    df_ml[col] = df_norm["_id_original"]
                elif col == "SKU":
                    df_ml[col] = df_norm["_id_original"]
                elif col == "ORIGINAL_PRICE":
                    df_ml[col] = df_norm["_preco_original"].round(2)
                elif col == "DISCOUNT_PERCENTAGE":
                    # Calcular percentual de desconto (ex: 0.05 -> 5)
                    df_ml[col] = int(round(desconto_percent * 100))
                elif col == "FINAL_PRICE":
                    df_ml[col] = df_norm["_preco_desconto"].round(2)
                elif col == "SUGGESTION":
                    df_ml[col] = ""
                elif col == "RECEIVES":
                    df_ml[col] = ""
                elif col == "LOYALTY_DISCOUNT_PERCENTAGE":
                    df_ml[col] = ""
                elif col == "LOYALTY_PRICE":
                    df_ml[col] = ""
                elif col == "LOYALTY_RECEIVES":
                    df_ml[col] = ""
                elif col == "STATUS":
                    df_ml[col] = "Ativo"
                elif col == "ACTION":
                    df_ml[col] = "Participar"
                elif col == "ERRORS":
                    df_ml[col] = ""
            
            return df_ml
        
        # Mapear para Shopee (padrão)
        df_marketplace = pd.DataFrame()
        
        # Mapear cada coluna do template para os dados normalizados
        for col in self.template["colunas_saida"]:
            if col == "ID do produto":
                df_marketplace[col] = df_norm["_id_original"]
            elif col == "Nome do Produto. (Opcional)":
                df_marketplace[col] = df_norm["_descricao_original"]
            elif col == "Nº de Ref. Parent SKU. (Opcional)":
                df_marketplace[col] = ""  # Vazio
            elif col == "ID de variação":
                df_marketplace[col] = df_norm["_id_original"]  # Mesmo ID do produto
            elif col == "Variação de nome. (Opcional)":
                df_marketplace[col] = ""  # Vazio
            elif col == "Nº de Ref. SKU. (Opcional)":
                df_marketplace[col] = df_norm["_id_original"]  # SKU/MLB original
            elif col == "Preço original (opcional)":
                df_marketplace[col] = df_norm["_preco_original"].round(2)
            elif col == "Preço de desconto":
                df_marketplace[col] = df_norm["_preco_desconto"]
            elif col == "Limite de compra (Opcional)":
                df_marketplace[col] = ""  # Vazio
        
        return df_marketplace
    
    def exportar_para_excel(self, df_marketplace, nome_sheet="Promoções"):
        """
        Exporta dados para Excel formatado profissionalmente
        
        Args:
            df_marketplace: DataFrame no formato do marketplace
            nome_sheet: Nome da aba Excel
            
        Returns:
            BytesIO com arquivo Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = nome_sheet
        
        # Escrever dados
        for r_idx, row in enumerate(dataframe_to_rows(df_marketplace, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Formatação do cabeçalho
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    # Alternância de cores
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    # Formatação de números (preços)
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                
                # Bordas
                thin_border = Border(
                    left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC")
                )
                cell.border = thin_border
        
        # Ajustar largura das colunas
        larguras = [18, 30, 20, 18, 25, 18, 18, 18, 20]
        for idx, largura in enumerate(larguras, 1):
            ws.column_dimensions[chr(64 + idx)].width = largura
        
        # Congelar primeira linha
        ws.freeze_panes = "A2"
        
        # Salvar em BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def calcular_impacto(self, df_marketplace, df_original):
        """
        Calcula o impacto econômico das promoções
        
        Args:
            df_marketplace: DataFrame no formato do marketplace
            df_original: DataFrame original com preços
            
        Returns:
            dict com métricas de impacto
        """
        preco_original = pd.to_numeric(df_marketplace["Preço original (opcional)"], errors='coerce')
        preco_desconto = pd.to_numeric(df_marketplace["Preço de desconto"], errors='coerce')
        
        economia_por_produto = preco_original - preco_desconto
        desconto_medio_percent = (economia_por_produto / preco_original * 100).mean()
        
        return {
            "economia_total": economia_por_produto.sum(),
            "economia_media": economia_por_produto.mean(),
            "desconto_medio_percent": desconto_medio_percent,
            "quantidade_produtos": len(df_marketplace),
        }
    
    def gerar_relatorio_impacto(self, df_marketplace, df_original):
        """
        Gera relatório detalhado do impacto das promoções
        
        Args:
            df_marketplace: DataFrame no formato do marketplace
            df_original: DataFrame original com preços
            
        Returns:
            dict com informações do relatório
        """
        impacto = self.calcular_impacto(df_marketplace, df_original)
        
        preco_original = pd.to_numeric(df_marketplace["Preço original (opcional)"], errors='coerce')
        preco_desconto = pd.to_numeric(df_marketplace["Preço de desconto"], errors='coerce')
        
        return {
            "total_produtos": len(df_marketplace),
            "economia_total": impacto["economia_total"],
            "economia_media_por_produto": impacto["economia_media"],
            "desconto_medio_percent": impacto["desconto_medio_percent"],
            "preco_medio_original": preco_original.mean(),
            "preco_medio_desconto": preco_desconto.mean(),
        }
