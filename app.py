"""
Aplicativo Streamlit para Precificação - Carblue
Novo fluxo: Relatório → Calculadora → Simulador
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO

from session_manager import inicializar_sessao, atualizar_margens
from pricing_calculator_v2 import PricingCalculatorV2
from price_simulator import PriceSimulator
from mercado_livre_processor import MercadoLivreProcessor
from promotion_exporter import PromotionExporter

# ============ FUNÇÕES DE FORMATAÇÃO ============
def formatar_moeda(valor):
    """Formata valor em reais no padrão brasileiro: R$ 1.234,50"""
    if pd.isna(valor) or valor is None:
        return "R$ 0,00"
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_percentual(valor):
    """Formata percentual no padrão brasileiro: 12,50%"""
    if pd.isna(valor) or valor is None:
        return "0,00%"
    return f"{valor:.2f}%".replace(".", ",")

def formatar_percentual_1casa(valor):
    """Formata percentual com 1 casa decimal: 12,5%"""
    if pd.isna(valor) or valor is None:
        return "0,0%"
    return f"{valor:.1f}%".replace(".", ",")

def formatar_excel_profissional(df, nome_sheet="Relatorio"):
    """Formata um DataFrame para Excel com estilos profissionais"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = nome_sheet
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="556B2F", end_color="556B2F", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0.00'
                
                thin_border = Border(
                    left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC")
                )
                cell.border = thin_border
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = "A2"
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    except ImportError:
        buffer = BytesIO()
        df.to_excel(buffer, index=False, sheet_name=nome_sheet)
        buffer.seek(0)
        return buffer

# Configurar página
st.set_page_config(
    page_title="Precificação Estratégica",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Inicializar sessão
inicializar_sessao()

# Estilos customizados
st.markdown("""
    <style>
    /* Tema Minimalista - Preto & Verde Militar */
    .stApp {
        background-color: #000000;
        color: #E8E8E8;
    }
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #ffffff;
        margin-bottom: 0.5rem;
        text-shadow: 0px 2px 4px rgba(0,0,0,0.3);
    }
    .section-header {
        font-size: 1.5rem;
        font-weight: bold;
        color: #ffffff;
        margin-top: 1rem;
        margin-bottom: 0.5rem;
    }
    .status-saudavel { color: #6B8E23; font-weight: bold; }
    .status-alerta { color: #fbbf24; font-weight: bold; }
    .status-prejuizo { color: #f87171; font-weight: bold; }
    
    .config-section {
        background-color: #1e293b;
        padding: 1rem;
        border-radius: 0.5rem;
        margin-bottom: 1rem;
        border: 1px solid #334155;
    }

    /* Estilização Premium Dark para o Sidebar */
    [data-testid="stSidebar"] {
        background-color: #000000;
        border-right: 1px solid #1e293b;
    }
    [data-testid="stSidebar"] .stMarkdown h1 {
        font-size: 1.2rem !important;
        font-weight: 700 !important;
        color: #ffffff;
        margin-bottom: 0.5rem !important;
    }
    [data-testid="stSidebar"] .stMarkdown p, 
    [data-testid="stSidebar"] .stMarkdown span,
    [data-testid="stSidebar"] label {
        font-size: 0.85rem !important;
        color: #cbd5e1;
    }
    [data-testid="stSidebar"] .stExpander {
        border: none !important;
        background-color: transparent !important;
        margin-bottom: 0.4rem !important;
    }
    [data-testid="stSidebar"] .stExpander details {
        border: 1px solid #1e293b !important;
        border-radius: 8px !important;
        background-color: #0f172a !important;
        transition: all 0.2s ease;
    }
    [data-testid="stSidebar"] .stExpander details:hover {
        border-color: #334155 !important;
        background-color: #1e293b !important;
    }
    [data-testid="stSidebar"] .stExpander summary p {
        font-size: 0.9rem !important;
        font-weight: 600 !important;
        color: #f8fafc !important;
    }
    /* Ajuste de inputs no sidebar para Dark Mode */
    [data-testid="stSidebar"] .stNumberInput input,
    [data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] {
        font-size: 0.8rem !important;
        background-color: #020617 !important;
        color: #ffffff !important;
        border-color: #1e293b !important;
    }
    [data-testid="stSidebar"] .stCaption {
        font-size: 0.75rem !important;
        color: #94a3b8 !important;
    }
    [data-testid="stSidebar"] hr {
        margin: 0.5rem 0 !important;
        border-color: #1e293b !important;
    }
    /* Estilo para Tabs no Dark Mode */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background-color: transparent;
    }
    .stTabs [data-baseweb="tab"] {
        height: 40px;
        white-space: pre;
        background-color: #1e293b;
        border-radius: 4px 4px 0px 0px;
        color: #94a3b8;
        padding: 10px 20px;
    }
    .stTabs [aria-selected="true"] {
        background-color: #334155 !important;
        color: #ffffff !important;
    }
    /* Remover tarja branca no topo */
    header[data-testid="stHeader"] {
        background-color: rgba(0,0,0,0) !important;
        color: #ffffff !important;
    }
    /* Corrigir legibilidade do File Uploader */
    [data-testid="stFileUploadDropzone"] {
        background-color: #0f172a !important;
        border: 1px dashed #334155 !important;
    }
    [data-testid="stFileUploadDropzone"] p, 
    [data-testid="stFileUploadDropzone"] span,
    [data-testid="stFileUploadDropzone"] small {
        color: #ffffff !important;
    }
    /* Corrigir botões de upload e outros botões secundários */
    .stButton button {
        background-color: #556B2F !important;
        color: #ffffff !important;
        border: 1px solid #556B2F !important;
    }
    .stButton button:hover {
        background-color: #6B8E23 !important;
        border-color: #6B8E23 !important;
        color: #ffffff !important;
    }
    
    /* PENTE FINO: Legibilidade Global */
    /* Forçar cor de texto em todos os parágrafos e spans */
    p, span, label, li, .stMarkdown {
        color: #fafafa !important;
    }
    /* Ajustar métricas */
    [data-testid="stMetricValue"] {
        color: #ffffff !important;
    }
    [data-testid="stMetricLabel"] p {
        color: #94a3b8 !important;
    }
    /* Ajustar tabelas e dataframes */
    .stDataFrame, [data-testid="stTable"] {
        background-color: #0f172a !important;
    }
    /* Ajustar legendas e captions */
    .stCaption, small {
        color: #94a3b8 !important;
    }
    /* Ajustar inputs globais (fora do sidebar também) */
    .stNumberInput input, .stTextInput input, .stSelectbox div[data-baseweb="select"] {
        background-color: #0f172a !important;
        color: #ffffff !important;
        border-color: #334155 !important;
    }
    /* Ajustar Sliders */
    .stSlider [data-baseweb="slider"] {
        background-color: transparent !important;
    }
    /* Ajustar Warnings e Infos */
    .stAlert {
        background-color: #1e293b !important;
        color: #ffffff !important;
        border: 1px solid #334155 !important;
    }
    /* Ajustar Divisores */
    hr {
        border-color: #334155 !important;
    }
    /* Forçar cor branca em títulos de expanders fora do sidebar */
    .stExpander summary p {
        color: #ffffff !important;
    }
    </style>
""", unsafe_allow_html=True)

# ============ SIDEBAR ============
st.sidebar.markdown("#  CONFIGURAÇÕES")
st.sidebar.markdown("---")

# 1. MARKETPLACES
with st.sidebar.expander("Marketplaces", expanded=False):
    for marketplace, config in st.session_state.marketplaces.items():
        st.markdown(f"**{marketplace}**")
        col1, col2 = st.columns([1, 1])
        
        with col1:
            comissao = st.number_input(
                "Comissão (%)",
                value=config["comissao"] * 100,
                min_value=0.0,
                max_value=100.0,
                step=0.1,
                key=f"comissao_{marketplace}",
            ) / 100
        
        with col2:
            taxa_fixa = st.number_input(
                "Taxa Fixa (R$)",
                value=config["custo_fixo"],
                min_value=0.0,
                step=0.1,
                key=f"taxa_fixa_{marketplace}",
            )
        
        st.session_state.marketplaces[marketplace]["comissao"] = comissao
        st.session_state.marketplaces[marketplace]["custo_fixo"] = taxa_fixa
        
        st.markdown("")
        st.divider()

# 2. REGIMES TRIBUTÁRIOS
with st.sidebar.expander("Regimes Tributários", expanded=False):
    for regime, config in st.session_state.regimes.items():
        st.markdown(f"**{regime}**")
        
        col1, col2, col3 = st.columns([1, 1, 1])
        
        with col1:
            ibs = st.number_input(
                "IBS (%)",
                value=config.get("ibs", 0.0) * 100,
                min_value=0.0,
                max_value=100.0,
                step=0.01,
                key=f"ibs_{regime}",
            ) / 100
        
        with col2:
            cbs = st.number_input(
                "CBS (%)",
                value=config.get("cbs", 0.0) * 100,
                min_value=0.0,
                max_value=100.0,
                step=0.01,
                key=f"cbs_{regime}",
            ) / 100
        
        with col3:
            impostos = st.number_input(
                "Impostos (%)",
                value=config.get("impostos_encargos", 0.0) * 100,
                min_value=0.0,
                max_value=100.0,
                step=0.1,
                key=f"impostos_{regime}",
            ) / 100
        
        st.session_state.regimes[regime]["ibs"] = ibs
        st.session_state.regimes[regime]["cbs"] = cbs
        st.session_state.regimes[regime]["impostos_encargos"] = impostos
        
        st.markdown("")
        st.divider()

# 3. MARGENS E PUBLICIDADE
with st.sidebar.expander("Margens e Publicidade", expanded=False):
    
    # Garantir que os valores sao validos
    valor_margem_bruta = float(st.session_state.get("margem_bruta_alvo", 30.0)) if st.session_state.get("margem_bruta_alvo") is not None else 30.0
    valor_margem_liquida = float(st.session_state.get("margem_liquida_minima", 10.0)) if st.session_state.get("margem_liquida_minima") is not None else 10.0
    valor_publicidade = float(st.session_state.get("percent_publicidade", 3.0)) if st.session_state.get("percent_publicidade") is not None else 3.0
    
    margem_bruta = st.slider(
        "Margem Bruta Alvo (%)",
        min_value=0.0,
        max_value=100.0,
        value=valor_margem_bruta,
        step=1.0,
        key="slider_margem_bruta"
    )
    st.caption(f"Margem Bruta: {formatar_percentual_1casa(margem_bruta)}")
    
    st.markdown("")
    
    margem_liquida = st.slider(
        "Margem Líquida Mínima (%)",
        min_value=0.0,
        max_value=100.0,
        value=valor_margem_liquida,
        step=1.0,
        key="slider_margem_liquida"
    )
    st.caption(f"Margem Líquida: {formatar_percentual_1casa(margem_liquida)}")
    
    st.markdown("")
    
    percent_pub = st.slider(
        "Percentual de Publicidade",
        min_value=0.0,
        max_value=100.0,
        value=valor_publicidade,
        step=0.1,
        key="slider_publicidade"
    )
    st.caption(f"Publicidade: {formatar_percentual_1casa(percent_pub)}")
    
    atualizar_margens(margem_bruta, margem_liquida, percent_pub)

# 4. CUSTOS OPERACIONAIS
with st.sidebar.expander("Custos Operacionais", expanded=False):
    
    custo_fixo_op = st.slider(
        "Custo Fixo Operacional (%)",
        min_value=0.0,
        max_value=100.0,
        value=st.session_state.get("custo_fixo_operacional", 0.0),
        step=0.1,
        key="slider_custo_fixo"
    )
    st.caption(f"Percentual da operação: {formatar_percentual_1casa(custo_fixo_op)}")
    
    st.markdown("")
    st.markdown("")
    
    taxa_devolucao = st.slider(
        "Taxa de Devoluções e Trocas (%)",
        min_value=0.0,
        max_value=100.0,
        value=st.session_state.get("taxa_devolucao", 0.0) * 100,
        step=0.1,
        key="slider_taxa_devolucao"
    ) / 100
    st.caption(f"Percentual de perdas: {formatar_percentual_1casa(taxa_devolucao * 100)}")
    
    st.session_state.custo_fixo_operacional = custo_fixo_op
    st.session_state.taxa_devolucao = taxa_devolucao
    
    st.markdown("")
    st.info("Os custos operacionais são aplicados como percentual do faturamento total.")

# 5. CARREGAR RELATÓRIO
with st.sidebar.expander("Carregar Relatório de Vendas", expanded=False):
    
    st.markdown("""
    **Formato esperado:**
    - **A:** SKU/MLB
    - **B:** Titulo
    - **C:** Custo Produto (R$)
    - **D:** Frete (R$)
    - **E:** Preco Atual (R$)
    - **F:** Tipo de Anuncio (opcional)
    - **G:** Quantidade Vendida (opcional - para Curva ABC)
    """)
    
    st.markdown("")
    
    # Criar e disponibilizar modelo de planilha
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    df_modelo = pd.DataFrame({
        'SKU/MLB': ['SKU001', 'SKU002', 'SKU003'],
        'Titulo': ['Produto Exemplo 1', 'Produto Exemplo 2', 'Produto Exemplo 3'],
        'Custo Produto (R$)': [25.50, 45.00, 15.75],
        'Frete (R$)': [8.00, 12.00, 5.00],
        'Preco Atual (R$)': [89.90, 149.90, 59.90],
        'Tipo de Anuncio': ['Classico', 'Premium', 'Classico'],
        'Quantidade Vendida': [150, 85, 320]
    })
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_modelo.to_excel(writer, index=False, sheet_name='Produtos')
        
        # Formatar planilha
        ws = writer.sheets['Produtos']
        header_fill = PatternFill(start_color="556B2F", end_color="556B2F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 20
    
    output.seek(0)
    
    st.download_button(
        label="Baixar Modelo de Planilha",
        data=output.getvalue(),
        file_name="modelo_precificacao.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    
    st.markdown("")
    
    uploaded_file = st.file_uploader(
        "Escolha um arquivo",
        type=["xlsx", "xls", "csv"],
        help="Relatório de vendas do Mercado Livre",
        key="sidebar_upload"
    )
    
    if uploaded_file is not None:
        try:
            with st.spinner("⏳ Processando..."):
                processor = MercadoLivreProcessor()
                
                if uploaded_file.name.endswith(".csv"):
                    df = processor.carregar_de_csv(uploaded_file)
                else:
                    df = processor.carregar_de_excel(uploaded_file)
                
                df_normalizado = processor.normalizar_relatorio_vendas(df)
                valido, mensagem = processor.validar_relatorio(df_normalizado)
                
                if valido:
                    df_agregado = processor.agregar_por_sku(df_normalizado)
                    st.session_state.relatorio_vendas = df_agregado
                    st.success(f" {len(df_agregado)} SKUs carregados com sucesso!")
                else:
                    st.error(f" {mensagem}")
        
        except Exception as e:
            st.error(f" Erro: {str(e)}")

# ============ ABAS PRINCIPAIS ============
st.markdown("---")
tab1, tab2, tab3, tab4, tab5 = st.tabs([" Home", "Calculadora de Precificação", "Simulador de Preço Alvo", " Dashboard", " Estratégias Promocionais"])

# ============ ABA 1: HOME ============
with tab1:
    # CSS personalizado para melhor visual
    st.markdown("""
    <style>
    .hero-section {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 40px 20px;
        border-radius: 15px;
        color: white;
        margin-bottom: 30px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    .hero-title {
        font-size: 2.5em;
        font-weight: 700;
        margin-bottom: 10px;
    }
    .hero-subtitle {
        font-size: 1.1em;
        opacity: 0.9;
    }
    .feature-card {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
        padding: 25px;
        border-radius: 12px;
        border-left: 5px solid #667eea;
        border: 1px solid rgba(255, 255, 255, 0.2);
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        margin-bottom: 20px;
        transition: all 0.3s ease;
        color: white;
    }
    .feature-card:hover {
        box-shadow: 0 4px 15px rgba(0,0,0,0.12);
        transform: translateY(-2px);
    }
    .feature-title {
        font-size: 1.3em;
        font-weight: 600;
        color: #667eea;
        margin-bottom: 10px;
    }
    .step-number {
        display: inline-block;
        background: #667eea;
        color: white;
        width: 40px;
        height: 40px;
        border-radius: 50%;
        text-align: center;
        line-height: 40px;
        font-weight: 700;
        margin-right: 15px;
        font-size: 1.1em;
    }
    .benefit-list {
        list-style: none;
        padding: 0;
    }
    .benefit-list li {
        padding: 12px 0;
        padding-left: 35px;
        position: relative;
        line-height: 1.6;
    }
    .benefit-list li:before {
        content: "✅";
        position: absolute;
        left: 0;
        color: #667eea;
        font-weight: bold;
        font-size: 1.2em;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Hero Section
    st.markdown("""
    <div class="hero-section">
        <div class="hero-title"> Precificação Estratégica</div>
        <div class="hero-subtitle">Análise Inteligente para Precificação Estratégica</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Resumo rápido em 3 colunas
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title">Análise Completa</div>
            <p>Cálculos precisos de custos, margens e rentabilidade em tempo real</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title">Oportunidades</div>
            <p>Identifique produtos com potencial de crescimento e ação</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title">Relatórios</div>
            <p>Exportação profissional em Excel com análises visuais</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Seção: Como Funciona
    st.markdown("## Como Funciona")
    
    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.markdown("""
        <div class="feature-card">
            <span class="step-number">1</span>
            <div class="feature-title">Configurar</div>
            <p>Defina marketplaces, regimes e margens na sidebar</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="feature-card">
            <span class="step-number">2</span>
            <div class="feature-title">Carregar</div>
            <p>Importe seu relatório de produtos</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="feature-card">
            <span class="step-number">3</span>
            <div class="feature-title">Calcular</div>
            <p>Análise automática de precificação</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="feature-card">
            <span class="step-number">4</span>
            <div class="feature-title">Simular</div>
            <p>Teste preços e margens</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="feature-card">
            <span class="step-number">5</span>
            <div class="feature-title">Analisar</div>
            <p>Veja oportunidades no dashboard</p>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        ### Funcionalidades Principais
        
        <ul class="benefit-list">
            <li><strong>Múltiplos Marketplaces:</strong> Mercado Livre, Shopee, Amazon e mais</li>
            <li><strong>Cálculos Precisos:</strong> Comissões, impostos, frete, publicidade</li>
            <li><strong>Análise ABC:</strong> Classificação automática por faturamento</li>
            <li><strong>Status de Saúde:</strong> Identifique produtos saudáveis, em alerta ou prejuízo</li>
            <li><strong>Oportunidades:</strong> Produtos B e C com potencial de crescimento</li>
            <li><strong>Simulador:</strong> Teste diferentes margens e preços</li>
            <li><strong>Relatórios:</strong> Exporte dados em Excel profissional</li>
            <li><strong>Dashboard:</strong> Visualizações interativas e análises detalhadas</li>
            <li><strong>Estratégias Promocionais:</strong> Gere templates de promoções para Shopee e Mercado Livre com descontos calculados</li>
        </ul>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Seção: Status de Saúde
    st.markdown("## Status de Saúde dos Produtos")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title">Saudável</div>
            <p><strong>Margem ≥ Alvo</strong></p>
            <p style="font-size: 0.9em; color: white;">Boa rentabilidade. Mantenha ou aumente volume.</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title">Alerta</div>
            <p><strong>Alvo > Margem ≥ Mínima</strong></p>
            <p style="font-size: 0.9em; color: white;">Rentabilidade aceitável. Monitore custos.</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title">Prejuízo</div>
            <p><strong>Margem < Mínima</strong></p>
            <p style="font-size: 0.9em; color: white;">Rentabilidade insuficiente. Revise preço.</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Seção: Classificação ABC
    st.markdown("## Classificação ABC")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title">Curva A</div>
            <p><strong>80% do faturamento</strong></p>
            <p style="font-size: 0.9em; color: white;">Produtos principais. Máxima atenção.</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title">Curva B</div>
            <p><strong>15% do faturamento</strong></p>
            <p style="font-size: 0.9em; color: white;">Produtos secundários. Oportunidades.</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title">Curva C</div>
            <p><strong>5% do faturamento</strong></p>
            <p style="font-size: 0.9em; color: white;">Produtos de baixa relevância.</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Dicas de Uso
    st.markdown("## Dicas para Melhor Aproveitar")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title"> Comece pela Configuração</div>
            <p>Dedique tempo para configurar corretamente os parâmetros de marketplaces e margens.</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title"> Use o Modelo</div>
            <p>Baixe e use o modelo de planilha fornecido para garantir dados corretos.</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title"> Analise o Dashboard</div>
            <p>Use o dashboard para identificar padrões e oportunidades de crescimento.</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="feature-card">
            <div class="feature-title"> Exporte Regularmente</div>
            <p>Mantenha histórico de análises para acompanhar tendências ao longo do tempo.</p>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    st.markdown("""
    <div style="text-align: center; padding: 20px; color: #888;">
        <p>Desenvolvido por Vinícius Lima</p>
    </div>
    """, unsafe_allow_html=True)

    # ============ ABA 2: CALCULADORA ============
with tab2:
    # CSS para Calculadora
    st.markdown("""
    <style>
    .calc-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 30px 20px;
        border-radius: 12px;
        color: white;
        margin-bottom: 30px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    .calc-title {
        font-size: 2em;
        font-weight: 700;
        margin-bottom: 5px;
    }
    .calc-subtitle {
        font-size: 0.95em;
        opacity: 0.9;
    }
    .input-card {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
        padding: 20px;
        border-radius: 12px;
        border: 1px solid rgba(255, 255, 255, 0.2);
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        margin-bottom: 20px;
    }
    .metric-card-calc {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
        padding: 20px;
        border-radius: 12px;
        border-top: 4px solid #667eea;
        border: 1px solid rgba(255, 255, 255, 0.2);
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        text-align: center;
    }
    .section-title-calc {
        color: white;
        font-size: 1.3em;
        font-weight: 700;
        color: white;
        margin: 30px 0 20px 0;
        padding-bottom: 10px;
        border-bottom: 3px solid #667eea;
    }
    .filter-card {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
        padding: 20px;
        border-radius: 12px;
        border: 1px solid rgba(255, 255, 255, 0.2);
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        margin-bottom: 20px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header
    st.markdown("""
    <div class="calc-header">
        <div class="calc-title">Calculadora de Precificação</div>
        <div class="calc-subtitle">Calcule preços, margens e rentabilidade de seus produtos</div>
    </div>
    """, unsafe_allow_html=True)
    
    if st.session_state.relatorio_vendas is None or st.session_state.relatorio_vendas.empty:
        st.info(" Carregue um relatório na sidebar para começar")
    else:
        st.markdown('<div class="section-title-calc">Configuração</div>', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            marketplace = st.selectbox(
                "Selecione o Marketplace",
                options=list(st.session_state.marketplaces.keys()),
                key="calc_marketplace"
            )
        
        with col2:
            regime = st.selectbox(
                "Selecione o Regime Tributário",
                options=list(st.session_state.regimes.keys()),
                key="calc_regime"
            )
        
        if st.button("Calcular Precificação", use_container_width=True, key="btn_calc"):
            try:
                calculator = PricingCalculatorV2(
                    marketplaces=st.session_state.marketplaces,
                    regimes=st.session_state.regimes,
                    margem_bruta_alvo=st.session_state.margem_bruta_alvo,
                    margem_liquida_minima=st.session_state.margem_liquida_minima,
                    percent_publicidade=st.session_state.get("percent_publicidade", 3.0),
                    custo_fixo_operacional=st.session_state.get("custo_fixo_operacional", 0.0),
                    taxa_devolucao=st.session_state.get("taxa_devolucao", 0.0),
                )
                
                df_resultado = calculator.calcular_dataframe(
                    st.session_state.relatorio_vendas,
                    marketplace,
                    regime
                )
                
                st.session_state.resultado_calculadora = df_resultado
                st.success("Cálculo realizado com sucesso!")
            
            except Exception as e:
                st.error(f"Erro ao calcular: {str(e)}")
        
        if "resultado_calculadora" in st.session_state:
            df_resultado = st.session_state.resultado_calculadora
            
            # Métricas
            st.markdown('<div class="section-title-calc">Resumo dos Resultados</div>', unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.markdown(f"""
                <div class="metric-card-calc">
                    <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Total de SKUs</div>
                    <div style="font-size: 2em; font-weight: 700; color: #667eea; margin: 10px 0;">{len(df_resultado)}</div>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                st.markdown(f"""
                <div class="metric-card-calc">
                    <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Margem Média</div>
                    <div style="font-size: 2em; font-weight: 700; color: #667eea; margin: 10px 0;">{formatar_percentual_1casa(df_resultado['Margem Bruta %'].mean())}</div>
                </div>
                """, unsafe_allow_html=True)
            with col3:
                st.markdown(f"""
                <div class="metric-card-calc">
                    <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Lucro Total</div>
                    <div style="font-size: 2em; font-weight: 700; color: #667eea; margin: 10px 0;">{formatar_moeda(df_resultado['Lucro R$'].sum())}</div>
                </div>
                """, unsafe_allow_html=True)
            with col4:
                saudaveis = len(df_resultado[df_resultado['Status'] == '🟢 Saudável'])
                st.markdown(f"""
                <div class="metric-card-calc" style="border-top-color: #22C55E;">
                    <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">🟢 Saudáveis</div>
                    <div style="font-size: 2em; font-weight: 700; color: #22C55E; margin: 10px 0;">{saudaveis}/{len(df_resultado)}</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("---")
            
            # Filtros
            st.markdown('<div class="section-title-calc">Filtros e Busca</div>', unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns([2, 2, 2, 2])
            
            with col1:
                pesquisa_sku = st.text_input("Pesquisar por SKU/MLB", placeholder="Digite SKU ou MLB")
            
            with col2:
                status_opcoes = ["Todos"] + list(df_resultado['Status'].unique())
                filtro_status = st.selectbox("Filtrar por Status", status_opcoes)
            
            with col3:
                if marketplace == "Mercado Livre" and "Tipo de Anuncio" in df_resultado.columns:
                    tipo_anuncio_opcoes = ["Todos"] + list(df_resultado['Tipo de Anuncio'].unique())
                    filtro_tipo_anuncio = st.selectbox("Filtrar por Tipo", tipo_anuncio_opcoes)
                else:
                    filtro_tipo_anuncio = "Todos"
            
            with col4:
                if "Curva ABC" in df_resultado.columns:
                    curva_abc_opcoes = ["Todos"] + sorted(df_resultado['Curva ABC'].unique())
                    filtro_curva_abc = st.selectbox("Filtrar por Curva ABC", curva_abc_opcoes)
                else:
                    filtro_curva_abc = "Todos"
            
            # Aplicar filtros
            df_filtrado = df_resultado.copy()
            
            if pesquisa_sku:
                col_sku = 'SKU ou MLB' if 'SKU ou MLB' in df_filtrado.columns else 'SKU'
                if col_sku in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado[col_sku].astype(str).str.contains(pesquisa_sku, case=False, na=False)]
            
            if filtro_status != "Todos":
                df_filtrado = df_filtrado[df_filtrado['Status'] == filtro_status]
            
            if marketplace == "Mercado Livre" and filtro_tipo_anuncio != "Todos" and "Tipo de Anuncio" in df_filtrado.columns:
                df_filtrado = df_filtrado[df_filtrado['Tipo de Anuncio'] == filtro_tipo_anuncio]
            
            if filtro_curva_abc != "Todos" and "Curva ABC" in df_filtrado.columns:
                df_filtrado = df_filtrado[df_filtrado['Curva ABC'] == filtro_curva_abc]
            
            st.markdown("---")
            
            # Tabela
            st.markdown(f'<div class="section-title-calc">Detalhes da Precificação ({len(df_filtrado)} produtos)</div>', unsafe_allow_html=True)
            st.dataframe(df_filtrado, use_container_width=True, hide_index=True)
            
            st.markdown("---")
            
            # Downloads
            st.markdown('<div class="section-title-calc">Downloads</div>', unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                if len(df_filtrado) > 0:
                    excel_filtrado = formatar_excel_profissional(df_filtrado, "Filtrado")
                    st.download_button(
                        label="Resultado Filtrado",
                        data=excel_filtrado,
                        file_name="calculadora_filtrado.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            with col2:
                df_saudaveis = df_resultado[df_resultado['Status'] == '🟢 Saudável']
                if len(df_saudaveis) > 0:
                    excel_saudaveis = formatar_excel_profissional(df_saudaveis, "🟢 Saudáveis")
                    st.download_button(
                        label=" 🟢 Saudáveis",
                        data=excel_saudaveis,
                        file_name="calculadora_saudaveis.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            with col3:
                df_alerta = df_resultado[df_resultado['Status'].str.contains('🟡 Alerta', na=False)]
                if len(df_alerta) > 0:
                    excel_alerta = formatar_excel_profissional(df_alerta, "Alerta")
                    st.download_button(
                        label=" 🟡 Em Alerta",
                        data=excel_alerta,
                        file_name="calculadora_alerta.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            with col4:
                df_prejuizo = df_resultado[df_resultado['Status'].str.contains('🔴 Prejuízo', na=False)]
                if len(df_prejuizo) > 0:
                    excel_prejuizo = formatar_excel_profissional(df_prejuizo, "Prejuízo")
                    st.download_button(
                        label=" 🔴 Em Prejuízo",
                        data=excel_prejuizo,
                        file_name="calculadora_prejuizo.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

# ============ ABA 3: SIMULADOR ============
with tab3:
    # CSS para Simulador
    st.markdown("""
    <style>
    .sim-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 30px 20px;
        border-radius: 12px;
        color: white;
        margin-bottom: 30px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    .sim-title {
        font-size: 2em;
        font-weight: 700;
        margin-bottom: 5px;
    }
    .sim-subtitle {
        font-size: 0.95em;
        opacity: 0.9;
    }
    .metric-card-sim {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
        padding: 20px;
        border-radius: 12px;
        border-top: 4px solid #667eea;
        border: 1px solid rgba(255, 255, 255, 0.2);
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        text-align: center;
    }
    .section-title-sim {
        color: white;
        font-size: 1.3em;
        font-weight: 700;
        color: white;
        margin: 30px 0 20px 0;
        padding-bottom: 10px;
        border-bottom: 3px solid #667eea;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header
    st.markdown("""
    <div class="sim-header">
        <div class="sim-title">Simulador de Preço Alvo</div>
        <div class="sim-subtitle">Simule diferentes preços e veja o impacto nas margens</div>
    </div>
    """, unsafe_allow_html=True)
    
    if st.session_state.relatorio_vendas is None or st.session_state.relatorio_vendas.empty:
        st.info(" Carregue um relatório na sidebar para começar")
    else:
        st.markdown('<div class="section-title-sim">Configuração</div>', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            marketplace = st.selectbox(
                "Selecione o Marketplace",
                options=list(st.session_state.marketplaces.keys()),
                key="sim_marketplace"
            )
        
        with col2:
            regime = st.selectbox(
                "Selecione o Regime Tributário",
                options=list(st.session_state.regimes.keys()),
                key="sim_regime"
            )
        
        if st.button("Simular Preços", use_container_width=True, key="btn_sim"):
            try:
                simulator = PriceSimulator(
                    marketplaces=st.session_state.marketplaces,
                    regimes=st.session_state.regimes,
                    margem_bruta_alvo=st.session_state.margem_bruta_alvo,
                    margem_liquida_minima=st.session_state.margem_liquida_minima,
                    percent_publicidade=st.session_state.get("percent_publicidade", 3.0),
                    custo_fixo_operacional=st.session_state.get("custo_fixo_operacional", 0.0),
                    taxa_devolucao=st.session_state.get("taxa_devolucao", 0.0),
                )
                
                df_simulacao = simulator.calcular_dataframe(
                    st.session_state.relatorio_vendas,
                    marketplace,
                    regime
                )
                
                st.session_state.resultado_simulador = df_simulacao
                st.success("Simulação realizada com sucesso!")
            
            except Exception as e:
                st.error(f"Erro ao simular: {str(e)}")
        
        if "resultado_simulador" in st.session_state:
            df_simulacao = st.session_state.resultado_simulador
            
            # Métricas
            st.markdown('<div class="section-title-sim">Resumo da Simulação</div>', unsafe_allow_html=True)
            
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.markdown(f"""
                <div class="metric-card-sim">
                    <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Preço Médio</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #667eea; margin: 10px 0;">{formatar_moeda(df_simulacao['Preço Sugerido'].mean())}</div>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                st.markdown(f"""
                <div class="metric-card-sim">
                    <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Preço Promo</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #667eea; margin: 10px 0;">{formatar_moeda(df_simulacao['Preço Promo Limite'].mean())}</div>
                </div>
                """, unsafe_allow_html=True)
            with col3:
                st.markdown(f"""
                <div class="metric-card-sim">
                    <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Lucro Bruto</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #667eea; margin: 10px 0;">{formatar_moeda(df_simulacao['Lucro Bruto'].sum())}</div>
                </div>
                """, unsafe_allow_html=True)
            with col4:
                st.markdown(f"""
                <div class="metric-card-sim">
                    <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Lucro Líquido</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #667eea; margin: 10px 0;">{formatar_moeda(df_simulacao['Lucro Líquido'].sum())}</div>
                </div>
                """, unsafe_allow_html=True)
            with col5:
                margem_media = df_simulacao['Margem Líquida %'].mean() if 'Margem Líquida %' in df_simulacao.columns else df_simulacao['Margem Bruta %'].mean()
                st.markdown(f"""
                <div class="metric-card-sim">
                    <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Margem Média</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #667eea; margin: 10px 0;">{formatar_percentual_1casa(margem_media)}</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("---")
            
            # Filtros
            st.markdown('<div class="section-title-sim">Filtros e Busca</div>', unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns([2, 2, 2, 2])
            
            with col1:
                pesquisa_sku = st.text_input("Pesquisar por SKU/MLB", placeholder="Digite SKU ou MLB", key="sim_pesquisa")
            
            with col2:
                status_opcoes = ["Todos"] + list(df_simulacao['Status'].unique()) if 'Status' in df_simulacao.columns else ["Todos"]
                filtro_status = st.selectbox("Filtrar por Status", status_opcoes, key="sim_status")
            
            with col3:
                if "Curva ABC" in df_simulacao.columns:
                    curva_abc_opcoes = ["Todos"] + sorted(df_simulacao['Curva ABC'].unique())
                    filtro_curva_abc = st.selectbox("Filtrar por Curva ABC", curva_abc_opcoes, key="sim_curva_abc")
                else:
                    filtro_curva_abc = "Todos"
            
            with col4:
                st.write("")
            
            # Aplicar filtros
            df_filtrado = df_simulacao.copy()
            
            if pesquisa_sku:
                col_sku = 'SKU ou MLB' if 'SKU ou MLB' in df_filtrado.columns else 'SKU'
                if col_sku in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado[col_sku].astype(str).str.contains(pesquisa_sku, case=False, na=False)]
            
            if 'Status' in df_filtrado.columns and filtro_status != "Todos":
                df_filtrado = df_filtrado[df_filtrado['Status'] == filtro_status]
            
            if filtro_curva_abc != "Todos" and "Curva ABC" in df_filtrado.columns:
                df_filtrado = df_filtrado[df_filtrado['Curva ABC'] == filtro_curva_abc]
            
            st.markdown("---")
            
            # Tabela
            st.markdown(f'<div class="section-title-sim"> Simulação de Preços ({len(df_filtrado)} produtos)</div>', unsafe_allow_html=True)
            st.dataframe(df_filtrado, use_container_width=True, hide_index=True)
            
            st.markdown("---")
            
            # Downloads
            st.markdown('<div class="section-title-sim">Downloads</div>', unsafe_allow_html=True)
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if len(df_filtrado) > 0:
                    excel_filtrado = formatar_excel_profissional(df_filtrado, "Simulação")
                    st.download_button(
                        label="Resultado Filtrado",
                        data=excel_filtrado,
                        file_name="simulador_filtrado.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            with col2:
                excel_completo = formatar_excel_profissional(df_simulacao, "Simulação Completa")
                st.download_button(
                    label="Simulação Completa",
                    data=excel_completo,
                    file_name="simulador_completo.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col3:
                st.write("")

# ============ ABA 4: DASHBOARD ============
with tab4:
    # CSS para Dashboard
    st.markdown("""
    <style>
    .dashboard-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 30px 20px;
        border-radius: 12px;
        color: white;
        margin-bottom: 30px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    .dashboard-title {
        font-size: 2em;
        font-weight: 700;
        margin-bottom: 5px;
    }
    .dashboard-subtitle {
        font-size: 0.95em;
        opacity: 0.9;
    }
    .metric-card {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
        padding: 20px;
        border-radius: 12px;
        border-top: 4px solid #667eea;
        border: 1px solid rgba(255, 255, 255, 0.2);
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        text-align: center;
    }
    .metric-value {
        font-size: 2em;
        font-weight: 700;
        color: #667eea;
        margin: 10px 0;
    }
    .metric-label {
        font-size: 0.9em;
        color: white;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .chart-container {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
        padding: 20px;
        border-radius: 12px;
        border: 1px solid rgba(255, 255, 255, 0.2);
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        margin-bottom: 20px;
    }
    .section-title {
        color: white;
        font-size: 1.3em;
        font-weight: 700;
        color: white;
        margin: 30px 0 20px 0;
        padding-bottom: 10px;
        border-bottom: 3px solid #667eea;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header do Dashboard
    st.markdown("""
    <div class="dashboard-header">
        <div class="dashboard-title">Dashboard de Análise</div>
        <div class="dashboard-subtitle">Visualize a saúde dos seus produtos e identifique oportunidades</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Obter dados da calculadora do session_state
    df_dashboard = st.session_state.resultado_calculadora if 'resultado_calculadora' in st.session_state else None
    
    if df_dashboard is None or len(df_dashboard) == 0:
        st.info("Carregue um relatório e calcule a precificação para visualizar o dashboard")
    else:
        try:
            import plotly.graph_objects as go
            
            # Contar produtos por status
            status_counts = df_dashboard['Status'].value_counts() if 'Status' in df_dashboard.columns else pd.Series()
            
            # Contar produtos por curva ABC
            curva_counts = df_dashboard['Curva ABC'].value_counts() if 'Curva ABC' in df_dashboard.columns else pd.Series()
            
            # Resumo em cards
            st.markdown('<div class="section-title">Resumo Geral</div>', unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-label">Total de Produtos</div>
                    <div class="metric-value">{len(df_dashboard)}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                saudaveis_total = len(df_dashboard[df_dashboard['Status'] == '🟢 Saudável']) if 'Status' in df_dashboard.columns else 0
                st.markdown(f"""
                <div class="metric-card" style="border-top-color: #22C55E;">
                    <div class="metric-label"> Saudável</div>
                    <div class="metric-value" style="color: #22C55E;">{saudaveis_total}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                alerta_total = len(df_dashboard[df_dashboard['Status'] == '🟡 Alerta']) if 'Status' in df_dashboard.columns else 0
                st.markdown(f"""
                <div class="metric-card" style="border-top-color: #EAB308;">
                    <div class="metric-label"> Alerta</div>
                    <div class="metric-value" style="color: #EAB308;">{alerta_total}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                prejuizo_total = len(df_dashboard[df_dashboard['Status'].astype(str).str.contains('🔴 Prejuízo', na=False)]) if 'Status' in df_dashboard.columns else 0
                st.markdown(f"""
                <div class="metric-card" style="border-top-color: #EF4444;">
                    <div class="metric-label"> Prejuízo</div>
                    <div class="metric-value" style="color: #EF4444;">{prejuizo_total}</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("---")
            
            # Gráficos em duas colunas
            st.markdown('<div class="section-title">Distribuição de Produtos</div>', unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                st.markdown("### Produtos por Status")
                if len(status_counts) > 0:
                    cores_status = {
                        '🟢 Saudável': '#22C55E',
                        '🟡 Alerta': '#EAB308',
                        '🔴 Prejuízo': '#EF4444',
                        ' Saudável': '#22C55E',
                        ' Alerta': '#EAB308',
                        ' Prejuízo': '#EF4444'
                    }
                    cores = [cores_status.get(str(status), '#999999') for status in status_counts.index]
                    
                    fig_status = go.Figure(data=[go.Pie(
                        labels=[str(label) for label in status_counts.index],
                        values=status_counts.values,
                        marker=dict(
                            colors=cores,
                            line=dict(color='rgba(255,255,255,0.3)', width=3)
                        ),
                        textposition='inside',
                        textinfo='label+percent',
                        textfont=dict(size=12, color='white', family='Arial Black'),
                        hovertemplate='<b>%{label}</b><br>%{value} produtos<br>%{percent}<extra></extra>',
                        pull=[0.05, 0.05, 0.05]
                    )])
                    fig_status.update_layout(
                        height=380,
                        margin=dict(l=20, r=20, t=20, b=20),
                        showlegend=False,
                        paper_bgcolor='rgba(0,0,0,0)',
                        plot_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='white', size=11),
                        hovermode='closest'
                    )
                    st.plotly_chart(fig_status, use_container_width=True)
                else:
                    st.info("Sem dados de status")
                st.markdown('</div>', unsafe_allow_html=True)
            
            with col2:
                st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                st.markdown("### Produtos por Curva ABC")
                if len(curva_counts) > 0:
                    cores_curva = {
                        'A': '#0066FF',
                        'B': '#FFFF99',
                        'C': '#FF9999',
                        'Curva A': '#0066FF',
                        'Curva B': '#FFFF99',
                        'Curva C': '#FF9999'
                    }
                    cores = [cores_curva.get(str(curva), '#999999') for curva in curva_counts.index]
                    
                    fig_curva = go.Figure(data=[go.Pie(
                        labels=[str(label) for label in curva_counts.index],
                        values=curva_counts.values,
                        marker=dict(
                            colors=cores,
                            line=dict(color='rgba(255,255,255,0.3)', width=3)
                        ),
                        textposition='inside',
                        textinfo='label+percent',
                        textfont=dict(size=12, color='white', family='Arial Black'),
                        hovertemplate='<b>%{label}</b><br>%{value} produtos<br>%{percent}<extra></extra>',
                        pull=[0.05, 0.05, 0.05]
                    )])
                    fig_curva.update_layout(
                        height=380,
                        margin=dict(l=20, r=20, t=20, b=20),
                        showlegend=False,
                        paper_bgcolor='rgba(0,0,0,0)',
                        plot_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='white', size=11),
                        hovermode='closest'
                    )
                    st.plotly_chart(fig_curva, use_container_width=True)
                else:
                    st.info("Sem dados de Curva ABC")
                st.markdown('</div>', unsafe_allow_html=True)

            
            st.markdown("---")
            
            # Análise Detalhada por Curva
            st.markdown('<div class="section-title">Análise Detalhada por Curva ABC</div>', unsafe_allow_html=True)
            
            if 'Curva ABC' in df_dashboard.columns and 'Status' in df_dashboard.columns:
                curvas_unicas = df_dashboard['Curva ABC'].unique()
                curvas_normalizadas = []
                for c in curvas_unicas:
                    if isinstance(c, str):
                        if c.startswith('Curva'):
                            curvas_normalizadas.append(c)
                        else:
                            curvas_normalizadas.append(f'Curva {c}')
                    else:
                        curvas_normalizadas.append(str(c))
                
                curvas = sorted(set(curvas_normalizadas))
                
                for curva in curvas:
                    curva_letra = curva.replace('Curva ', '').strip()
                    
                    df_curva = df_dashboard[df_dashboard['Curva ABC'].astype(str).str.contains(curva_letra, na=False)]
                    
                    saudaveis_curva = len(df_curva[df_curva['Status'] == '🟢 Saudável'])
                    alerta_curva = len(df_curva[df_curva['Status'] == '🟡 Alerta'])
                    prejuizo_curva = len(df_curva[df_curva['Status'].astype(str).str.contains('🔴 Prejuízo', na=False)])
                    
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.markdown(f"""
                        <div class="metric-card">
                            <div class="metric-label">Curva {curva_letra}</div>
                            <div class="metric-value">{len(df_curva)}</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col2:
                        st.markdown(f"""
                        <div class="metric-card" style="border-top-color: #22C55E;">
                            <div class="metric-label"> Saudável</div>
                            <div class="metric-value" style="color: #22C55E;">{saudaveis_curva}</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col3:
                        st.markdown(f"""
                        <div class="metric-card" style="border-top-color: #EAB308;">
                            <div class="metric-label"> Alerta</div>
                            <div class="metric-value" style="color: #EAB308;">{alerta_curva}</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col4:
                        st.markdown(f"""
                        <div class="metric-card" style="border-top-color: #EF4444;">
                            <div class="metric-label"> Prejuízo</div>
                            <div class="metric-value" style="color: #EF4444;">{prejuizo_curva}</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    st.markdown("")
            else:
                st.info("Colunas 'Curva ABC' ou 'Status' não encontradas nos dados")
            
            st.markdown("---")
            
            # Oportunidades de Ação
            st.markdown('<div class="section-title">Oportunidades de Ação</div>', unsafe_allow_html=True)
            
            # Informativo sobre critérios
            st.markdown("""
            <div style="background: rgba(102, 126, 234, 0.1); border-left: 4px solid #667eea; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                <div style="color: white; font-size: 0.95em; line-height: 1.6;">
                    <strong>Sobre estas Oportunidades:</strong><br>
                    Produtos exibidos abaixo são da <strong>Curva B e C</strong> com <strong>margem bruta 5% ou mais acima</strong> da margem alvo configurada. 
                    Estes produtos têm potencial para ações diferenciadas como promoções, aumentos de volume ou investimentos em marketing.
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            oportunidades = df_dashboard[
                (df_dashboard['Curva ABC'].astype(str).str.contains('B', na=False) | df_dashboard['Curva ABC'].astype(str).str.contains('C', na=False)) &
                (df_dashboard['Status'] == '🟢 Saudável')
            ]
            
            # Armazenar oportunidades em session_state para uso na aba de Estrategias Promocionais
            st.session_state.lista_oportunidades = oportunidades.copy()
            
            if len(oportunidades) > 0:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #d4edda 0%, #c3e6cb 100%); padding: 20px; border-radius: 12px; margin-bottom: 20px; border-left: 5px solid #22C55E;">
                    <div style="font-size: 1.2em; font-weight: 700; color: #155724;"> {len(oportunidades)} Oportunidades Encontradas</div>
                    <div style="color: #155724; margin-top: 5px;">Produtos da Curva B e C com margens saudáveis - Potencial para ações diferenciadas</div>
                </div>
                """, unsafe_allow_html=True)
                
                # Colunas para exibir
                colunas_oportunidade = []
                if 'SKU ou MLB' in oportunidades.columns:
                    colunas_oportunidade.append('SKU ou MLB')
                if 'Titulo' in oportunidades.columns:
                    colunas_oportunidade.append('Titulo')
                if 'Curva ABC' in oportunidades.columns:
                    colunas_oportunidade.append('Curva ABC')
                if 'Preco Atual (R$)' in oportunidades.columns:
                    colunas_oportunidade.append('Preco Atual (R$)')
                if 'Lucro R$' in oportunidades.columns:
                    colunas_oportunidade.append('Lucro R$')
                if 'Margem Bruta %' in oportunidades.columns:
                    colunas_oportunidade.append('Margem Bruta %')
                if 'Status' in oportunidades.columns:
                    colunas_oportunidade.append('Status')
                
                if colunas_oportunidade:
                    st.dataframe(
                        oportunidades[colunas_oportunidade].reset_index(drop=True),
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    # Botão para baixar oportunidades
                    buffer = BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        oportunidades[colunas_oportunidade].to_excel(writer, sheet_name='Oportunidades', index=False)
                    
                    buffer.seek(0)
                    st.download_button(
                        label="Baixar Oportunidades em Excel",
                        data=buffer,
                        file_name=f"oportunidades_curva_bc.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_download_oportunidades"
                    )
            else:
                st.markdown("""
                <div style="background: linear-gradient(135deg, #fff3cd 0%, #ffeaa7 100%); padding: 20px; border-radius: 12px; border-left: 5px solid #EAB308;">
                    <div style="font-size: 1.1em; font-weight: 700; color: #856404;"> Nenhuma Oportunidade Encontrada</div>
                    <div style="color: #856404; margin-top: 5px;">Todos os produtos Curva B e C estão em Alerta ou Prejuízo. Revise suas estratégias de preço.</div>
                </div>
                """, unsafe_allow_html=True)
        
        except Exception as e:
            st.error(f"Erro ao gerar dashboard: {str(e)}")



# ============ ABA 5: ESTRATÉGIAS PROMOCIONAIS ============
with tab5:
    # CSS para Estratégias Promocionais
    st.markdown("""
    <style>
    .promo-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 30px 20px;
        border-radius: 12px;
        color: white;
        margin-bottom: 30px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    .promo-title {
        font-size: 2em;
        font-weight: 700;
        margin-bottom: 5px;
    }
    .promo-subtitle {
        font-size: 0.95em;
        opacity: 0.9;
    }
    .section-title-promo {
        color: white;
        font-size: 1.3em;
        font-weight: 700;
        margin: 30px 0 20px 0;
        padding-bottom: 10px;
        border-bottom: 3px solid #667eea;
    }
    .metric-card-promo {
        background: rgba(102, 126, 234, 0.1);
        backdrop-filter: blur(10px);
        padding: 20px;
        border-radius: 12px;
        border-top: 4px solid #667eea;
        border: 1px solid rgba(102, 126, 234, 0.2);
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        text-align: center;
    }
    .step-container {
        background: rgba(255, 255, 255, 0.05);
        border: 1px solid rgba(102, 126, 234, 0.3);
        border-radius: 12px;
        padding: 25px;
        margin-bottom: 25px;
        backdrop-filter: blur(10px);
    }
    .step-number {
        display: inline-block;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        width: 40px;
        height: 40px;
        border-radius: 50%;
        text-align: center;
        line-height: 40px;
        font-weight: 700;
        font-size: 1.1em;
        margin-right: 15px;
    }
    .step-title {
        display: inline-block;
        color: white;
        font-size: 1.1em;
        font-weight: 600;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header
    st.markdown("""
    <div class="promo-header">
        <div class="promo-title"> Estratégias Promocionais</div>
        <div class="promo-subtitle">Configure descontos por categoria e exporte para o marketplace selecionado</div>
    </div>
    """, unsafe_allow_html=True)
    
    if st.session_state.relatorio_vendas is None or st.session_state.relatorio_vendas.empty:
        st.info("Carregue um relatório na sidebar para começar")
    else:
        # Seção 1: Seleção de Marketplace
        st.markdown("""
        <div class="step-container">
            <span class="step-number">1</span>
            <span class="step-title">Selecione o Marketplace</span>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            marketplace_selecionado = st.selectbox(
                "Marketplace",
                options=["Shopee", "Mercado Livre"],
                key="promo_marketplace"
            )
        
        with col2:
            st.markdown("")
            st.markdown("")
            st.info(f"📦 Canal: **{marketplace_selecionado}**")
        
        # Seção 2: Seleção de Categoria
        st.markdown("""
        <div class="step-container">
            <span class="step-number">2</span>
            <span class="step-title">Selecione a Categoria de Produtos</span>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            tipo_filtro = st.radio(
                "Tipo de Filtro",
                options=["Oportunidades", "Curva ABC"],
                key="promo_tipo_filtro"
            )
        
        with col2:
            if tipo_filtro == "Oportunidades":
                categoria_selecionada = "Oportunidades"
                st.markdown("")
                st.markdown("")
                st.info("Produtos Curva B/C com margem saudável")
            else:
                categoria_selecionada = st.selectbox(
                    "Selecione a Curva",
                    options=["Curva A", "Curva B", "Curva C"],
                    key="promo_curva"
                )
        
        # Mapeamento de categorias
        categoria_map = {
            "Oportunidades": "oportunidade",
            "Curva A": "curva_a",
            "Curva B": "curva_b",
            "Curva C": "curva_c",
        }
        categoria_filtro = categoria_map[categoria_selecionada]
        
        # Seção 3: Configuração de Desconto e Margem
        st.markdown("""
        <div class="step-container">
            <span class="step-number">3</span>
            <span class="step-title">Configure o Desconto e Critérios</span>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            desconto_percent = st.slider(
                "Percentual de Desconto (%)",
                min_value=0.0,
                max_value=50.0,
                value=5.0,
                step=0.5,
                key="promo_desconto"
            ) / 100
        
        with col2:
            margem_minima = st.number_input(
                "Margem Mínima (%)",
                min_value=0.0,
                max_value=100.0,
                value=15.0,
                step=1.0,
                key="promo_margem_minima"
            )
        
        with col3:
            st.markdown("")
            st.markdown("")
            st.info(f"💰 Desconto: **{desconto_percent*100:.1f}%**")
        
        # Botão para processar
        if st.button("Processar e Visualizar", use_container_width=True, key="btn_promo_processar"):
            try:
                with st.spinner("⏳ Processando dados..."):
                    # Inicializar exportador com o marketplace selecionado
                    exporter = PromotionExporter(marketplace=marketplace_selecionado)
                    
                    # Se for oportunidade, usar a lista ja calculada no Dashboard
                    if categoria_filtro == "oportunidade":
                        df_filtrado = st.session_state.get("lista_oportunidades", None)
                        if df_filtrado is None or len(df_filtrado) == 0:
                            st.error("Erro: Nenhuma oportunidade encontrada. Verifique o Dashboard.")
                            st.stop()
                    else:
                        # Para Curva ABC, usar resultado_calculadora
                        df_base = st.session_state.get("resultado_calculadora", None)
                        if df_base is None:
                            st.error("Erro: Nenhum dado de Dashboard disponivel.")
                            st.stop()
                        
                        df_filtrado = exporter.filtrar_por_categoria(
                            df_base,
                            categoria=categoria_filtro,
                            margem_minima=margem_minima,
                            margem_alvo=st.session_state.get("slider_margem_bruta", 30.0)
                        )
                    
                    if len(df_filtrado) == 0:
                        st.warning(f"⚠️ Nenhum produto encontrado na categoria '{categoria_selecionada}'")
                    else:
                        # Mapear para marketplace
                        df_marketplace = exporter.mapear_dados_para_marketplace(df_filtrado, desconto_percent=desconto_percent)
                        
                        # Armazenar em session_state
                        st.session_state.df_marketplace_processado = df_marketplace
                        st.session_state.df_marketplace_original = df_filtrado
                        st.session_state.marketplace_ativo = marketplace_selecionado
                        
                        # Calcular impacto
                        relatorio = exporter.gerar_relatorio_impacto(df_marketplace, df_filtrado)
                        
                        # Exibir métricas
                        st.markdown('<div class="section-title-promo">4. Resumo do Impacto</div>', unsafe_allow_html=True)
                        
                        col1, col2, col3, col4, col5 = st.columns(5)
                        
                        with col1:
                            st.markdown(f"""
                            <div class="metric-card-promo">
                                <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Produtos</div>
                                <div style="font-size: 1.8em; font-weight: 700; color: #667eea; margin: 10px 0;">{relatorio['total_produtos']}</div>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with col2:
                            st.markdown(f"""
                            <div class="metric-card-promo">
                                <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Economia Total</div>
                                <div style="font-size: 1.8em; font-weight: 700; color: #667eea; margin: 10px 0;">{formatar_moeda(relatorio['economia_total'])}</div>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with col3:
                            st.markdown(f"""
                            <div class="metric-card-promo">
                                <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Economia Média</div>
                                <div style="font-size: 1.8em; font-weight: 700; color: #667eea; margin: 10px 0;">{formatar_moeda(relatorio['economia_media_por_produto'])}</div>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with col4:
                            st.markdown(f"""
                            <div class="metric-card-promo">
                                <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Preço Médio</div>
                                <div style="font-size: 1.8em; font-weight: 700; color: #667eea; margin: 10px 0;">{formatar_moeda(relatorio['preco_medio_original'])}</div>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with col5:
                            st.markdown(f"""
                            <div class="metric-card-promo">
                                <div style="font-size: 0.9em; color: white; text-transform: uppercase; letter-spacing: 0.5px;">Desconto Médio</div>
                                <div style="font-size: 1.8em; font-weight: 700; color: #667eea; margin: 10px 0;">{formatar_percentual_1casa(relatorio['desconto_medio_percent'])}</div>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        st.markdown("---")
                        
                        # Exibir tabela
                        st.markdown('<div class="section-title-promo">5. Produtos Selecionados</div>', unsafe_allow_html=True)
                        st.dataframe(df_marketplace, use_container_width=True, hide_index=True)
                        
                        st.markdown("---")
                        
                        # Download
                        st.markdown('<div class="section-title-promo">6. Download da Planilha</div>', unsafe_allow_html=True)
                        
                        buffer = exporter.exportar_para_excel(
                            df_marketplace,
                            f"Promoção {marketplace_selecionado} - {categoria_selecionada}"
                        )
                        
                        nome_arquivo = f"{marketplace_selecionado.lower()}_promocoes_{categoria_filtro}_{int(desconto_percent*100)}pct.xlsx"
                        
                        st.download_button(
                            label=f"📥 Baixar Planilha {marketplace_selecionado} ({len(df_marketplace)} produtos)",
                            data=buffer,
                            file_name=nome_arquivo,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        
                        st.success(f"✅ Planilha pronta para upload no {marketplace_selecionado}!")
                        
            except Exception as e:
                st.error(f"❌ Erro ao processar: {str(e)}")
# Forçar recarregamento do Streamlit - Mon Feb  9 14:05:19 EST 2026
